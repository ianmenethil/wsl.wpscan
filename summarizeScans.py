import os
import datetime
import logging
import re
import shutil
from typing import Any
import csv
from rich.logging import RichHandler
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO,
                    handlers=[
                        RichHandler(show_time=True,
                                    omit_repeated_times=False,
                                    show_level=True,
                                    show_path=True,
                                    enable_link_path=True,
                                    markup=True,
                                    rich_tracebacks=True,
                                    tracebacks_width=200,
                                    tracebacks_show_locals=False,
                                    tracebacks_theme='monokai',
                                    tracebacks_extra_lines=0,
                                    log_time_format='[%X]')
                    ])


def readTxtFile(input_file) -> str:
    """Reads a text file and returns its content."""
    with open(input_file, 'r', encoding='utf-8') as file:
        return file.read()


def apply_formatting(worksheet) -> None:
    """Apply formatting to a worksheet."""
    purple_fill = PatternFill(start_color="9370DB", end_color="9370DB", fill_type="solid")
    light_purple_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    bold_black_font = Font(name="Open Sans", color="000000", bold=True)
    bold_red_font = Font(name="Open Sans", color="FF0000", bold=True)
    open_sans_font = Font(name="Open Sans", size=10)
    for cell in worksheet[1]:
        cell.fill = purple_fill
        cell.font = bold_black_font
        cell.font = Font(bold=True)  # Make the header row bold
    for cell in worksheet['C']:
        cell.font = bold_red_font
    # Set font color for each row in the vulnerability column to black
    for cell in worksheet['D']:
        cell.font = bold_black_font
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = open_sans_font
    # Apply different color rotation to each row
    for i, row in enumerate(worksheet.iter_rows(min_row=2), start=1):
        row_fill = light_purple_fill if i % 2 == 0 else purple_fill
        for cell in row:
            cell.fill = row_fill


def getUniqueFilename(base_filename) -> str:
    """Get a unique filename by appending a date and a number if the file already exists."""
    filename, extension = os.path.splitext(base_filename)
    counter = 1
    while os.path.exists(filename):
        filename = f"{filename}_{counter}{extension}"
        counter += 1
    return filename


def getWorkbook(filename):
    """Get a workbook object. If the file exists, load the existing data. Otherwise, create a new workbook."""
    if os.path.exists(filename):
        return openpyxl.load_workbook(filename)
    return openpyxl.Workbook()


def sanitize_string(value) -> Any:
    """Remove or replace characters that cannot be used in Excel."""
    illegal_chars = [
        '\x00',
        '\x01',
        '\x02',
        '\x03',
        '\x04',
        '\x05',
        '\x06',
        '\x07',
        '\x08',
        '\x0b',
        '\x0c',
        '\x0d',
        '\x0e',
        '\x0f',
        '\x10',
        '\x11',
        '\x12',  # pylint: disable=line-too-long
        '\x13',
        '\x14',
        '\x15',
        '\x16',
        '\x17',
        '\x18',
        '\x19',
        '\x1a',
        '\x1b',
        '\x1c',
        '\x1d',
        '\x1e',
        '\x1f',
        '\x7f',
        '\x80',
        '\x81',
        '\x82',
        '\x83',
        '\x84',
        '\x85',
        '\x86',
        '\x87',
        '\x88',
        '\x89',
        '\x8a',
        '\x8b',
        '\x8c',
        '\x8d',
        '\x8e',
        '\x8f',
        '\x90',
        '\x91',
        '\x92',
        '\x93',
        '\x94',
        '\x95',
        '\x96',
        '\x97',
        '\x98',
        '\x99',
        '\x9a',
        '\x9b',
        '\x9c',
        '\x9d',
        '\x9e',
        '\x9f'
    ]
    for char in illegal_chars:
        value = value.replace(char, "")  # Remove the illegal character
    return value


def getAllTxtInDir(indir) -> list[Any]:
    """Find all .txt files within the specified directory and return their paths as a list."""
    txt_files = []
    for filename in os.listdir(indir):
        if filename.endswith('.txt'):
            txt_files.append(os.path.join(indir, filename))
    return txt_files


def delAll(directory, filename_substr) -> bool:
    deleted = False
    for foldername, _, filenames in os.walk(directory):
        for filename in filenames:
            if filename_substr in filename:
                file_path = os.path.join(foldername, filename)
                os.remove(file_path)
                logger.info(f"Deleted {file_path}")
                deleted = True
    return deleted


def copyRecurse(source_dir, destination_dir, extension='.txt') -> bool:
    """Copy all files from a directory recursively."""
    try:
        if os.path.isdir(source_dir):
            if not os.path.exists(destination_dir):
                os.makedirs(destination_dir)
            os.listdir(destination_dir)
            files = os.listdir(source_dir)
            for file in files:
                src_file = os.path.join(source_dir, file)
                dest_file = os.path.join(destination_dir, file)
                if os.path.isdir(src_file):
                    copyRecurse(src_file, dest_file, extension)  # Call the function recursively
                else:
                    if file.endswith(extension):
                        shutil.copyfile(src_file, dest_file)
            # files_after_copy = os.listdir(destination_dir)  # List files after copy
            os.listdir(destination_dir)
        else:
            if source_dir.endswith(extension):
                shutil.copyfile(source_dir, destination_dir)
            return True
    except Exception as e:
        logger.error(f'e in copy {e}')
        return False
    return True


def copyCsvFiles(source_dir, destination_dir) -> bool:
    """Copy all CSV files from a directory recursively to the destination directory."""
    try:
        if os.path.isdir(source_dir):
            files = os.listdir(source_dir)
            for file in files:
                src_file = os.path.join(source_dir, file)
                if os.path.isdir(src_file):
                    copyCsvFiles(src_file, destination_dir)  # Call the function recursively
                else:
                    if file.endswith('.csv'):
                        dest_file = os.path.join(destination_dir, file)
                        shutil.copyfile(src_file, dest_file)
                        # logger.info(f'Copying CSV files {src_file} to {dest_file}')
        else:
            if source_dir.endswith('.csv'):
                shutil.copyfile(source_dir, destination_dir)
            return True
    except Exception as e:
        logger.error(f'Error in copying CSV files: {e}')
        return False
    return False


def moveLogsToBackup(sourceDir, destDir) -> None:
    try:
        if os.path.isdir(sourceDir):
            files = os.listdir(sourceDir)
            for file in files:
                src_file = os.path.join(sourceDir, file)
                dest_file = os.path.join(destDir, file)
                if file.endswith('.txt') or file.endswith('.csv') or file.endswith('.log') and src_file != dest_file:
                    shutil.move(src_file, dest_file)
    except Exception as e:
        logger.error(e)


def setupDirs(OUTDIR, TEMPDIR, BACKUPDIR, TWISTERDIR, LOGSBACKUPDIR, DNSTWISTDIR) -> None:
    """ Ensure necessary directories exist. """
    try:

        def create_directory(directory) -> None:
            if not os.path.exists(directory):
                os.makedirs(directory)
                # logger.info(f"Created directory: {directory}")
            else:
                pass
                # logger.warning(f"Directory already exists: {directory}")

        create_directory(OUTDIR)
        create_directory(TEMPDIR)
        create_directory(BACKUPDIR)
        create_directory(TWISTERDIR)
        create_directory(LOGSBACKUPDIR)
        create_directory(DNSTWISTDIR)
    except Exception as e:
        logger.error(f'Error in setting up directories: {e}')


def cleanLogsHelper(file_path, start_line) -> Any:
    """ Process a single file and delete all lines before a specific update line. """
    end_line = "DEBUG - Parsing WPScan output"
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()
        start_pattern = re.compile(re.escape(start_line))
        start_index = None
        for i, line in enumerate(lines):
            if start_pattern.search(line):
                start_index = i
                break
        end_pattern = re.compile(re.escape(end_line))
        end_index = None
        for i, line in enumerate(lines[start_index or 0:], start_index or 0):
            if end_pattern.search(line):
                end_index = i
                break
        if start_index is not None and end_index is not None:
            lines = lines[start_index:end_index]
        elif start_index is not None:
            lines = lines[start_index:]
        new_file_path = file_path.rsplit('.', 1)[0] + '.txt'
        with open(new_file_path, 'w', encoding='utf-8') as file:
            file.writelines(lines)
    except Exception as e:
        logger.error(f"Error processing file {file_path}: {e}")


def cleanLogs(directory, update_line="[i] Updating the Database ...") -> None:
    """ Delete lines before a specific update line in all files within the directory. """
    logger.info(f"Deleting lines before {update_line} in all files within {directory}")
    if not os.path.isdir(directory):
        logger.error(f"{directory} is not a directory")
        return
    for filename in os.listdir(directory):
        if filename.endswith('.log'):
            file_path = os.path.join(directory, filename)
            cleanLogsHelper(file_path, update_line)


section_start_interesting_findings = 'Interesting Finding(s):'
section_end_interesting_findings = '[i] The main theme could not be detected.'
section_start_vuln_plugins = r"\[\+\] Enumerating Vulnerable Plugins.*"
section_end_vuln_plugins = '[i] No plugins Found.' or '[+] Enumerating Vulnerable Themes (via Aggressive Methods)'  # pylint: disable=line-too-long
section_start_vuln_themes = '[+] Enumerating Vulnerable Themes (via Aggressive Methods)'
section_end_vuln_themes = r"\[\+\] Enumerating Timthumbs.*"
section_start_timthumbs = '[+] Enumerating Timthumbs (via Aggressive Methods)'
section_end_timthumbs = '[i] No Timthumbs Found.' or '[+] Enumerating Config Backups (via Aggressive Methods)'  # pylint: disable=line-too-long
section_start_config_backups = '[+] Enumerating Config Backups (via Aggressive Methods)'
section_end_config_backups = '[i] No Config Backups Found.' or '[+] Enumerating DB Exports (via Aggressive Methods)'  # pylint: disable=line-too-long
section_start_db_exports = '[+] Enumerating DB Exports (via Aggressive Methods)'
section_end_db_exports = '[i] No DB Exports Found.' or '[+] Enumerating Users (via Aggressive Methods)'  # pylint: disable=line-too-long
section_start_users = '[+] Enumerating Users (via Aggressive Methods)'
section_end_users = '[i] User(s) Identified:' or '[+] WPScan DB API OK'  # WPScan DB API OK means no users were identified


def processFileForResults(txt_filename) -> list[Any]:
    """Process a text file to extract plugin details, vulnerabilities, and interesting findings."""
    content = readTxtFile(txt_filename)
    results = []
    lines = content.split('\n')
    current_item = ''  # This will hold either a plugin or theme name
    current_details = ''
    current_vuln_count = ''
    in_interesting_findings = False
    in_plugins_section = False
    in_themes_section = False  # Flag to check if we are in the themes section
    interesting_findings = ''

    for _, line in enumerate(lines):
        line = line.strip()
        if in_interesting_findings:
            if line in [section_end_interesting_findings, '[i]']:
                in_interesting_findings = False
                continue
            interesting_findings += line + '\n'
        elif in_plugins_section or in_themes_section:
            if line.startswith('| Location:'):
                # Save previous plugin/theme details, if any
                if current_item:
                    results.append({
                        'filename': txt_filename,
                        'interesting_findings': interesting_findings.strip(),
                        'plugin/theme': current_item,  # Use 'item' to generalize plugins and themes
                        'vuln_count': current_vuln_count,
                        'details': current_details.strip()
                    })
                    interesting_findings = ''  # Reset interesting findings for the next item
                # Capture the plugin/theme name
                current_item = line.split('/')[-2]
                current_details = line + '\n'  # Start accumulating the details
                # current_details = line  # Start accumulating the details
                current_vuln_count = ''  # Reset the vulnerability count for the new item
            elif '| [!]' in line:
                # Vulnerability detail line
                current_details += line + '\n'  # Accumulate details
                # current_details += line  # Accumulate details
                if 'vulnerabilities identified:' in line or 'vulnerability identified:' in line:
                    current_vuln_count = line.split(' ')[2]  # Capture vulnerability count
            elif line in [section_end_vuln_plugins, section_start_vuln_themes]:
                # Check if we've reached the end of plugins or start of themes
                in_plugins_section = False
                in_themes_section = line == section_start_vuln_themes
                if current_item:
                    # Save the last item's details
                    results.append({
                        'filename': txt_filename,
                        'interesting_findings': interesting_findings.strip(),
                        'plugin/theme': current_item,
                        'vuln_count': current_vuln_count,
                        'details': current_details.strip()
                    })
                    interesting_findings = ''  # Reset for the next section
                    current_item = ''
                    current_details = ''
                    current_vuln_count = ''
            elif re.match(section_end_vuln_themes, line):
                # End of themes section
                in_themes_section = False
                if current_item:
                    # Save the last theme's details
                    results.append({
                        'filename': txt_filename,
                        'interesting_findings': interesting_findings.strip(),
                        'plugin/theme': current_item,
                        'vuln_count': current_vuln_count,
                        'details': current_details.strip()
                    })
                    interesting_findings = ''  # Reset for the next section
                    current_item = ''
                    current_details = ''
                    current_vuln_count = ''
            else:
                current_details += line + '\n'  # Continue accumulating details
        elif line == section_start_interesting_findings:
            in_interesting_findings = True
        elif re.search(section_start_vuln_plugins, line) or re.search('[i] Plugin(s) Identified:', line):
            in_plugins_section = True
            in_themes_section = False
        elif re.search(section_start_vuln_themes, line) or re.search('[i] Theme(s) Identified:', line):
            in_plugins_section = True
            in_themes_section = False
    # Capture the last item after the loop ends, if any
    if current_item and current_details.strip():
        results.append({
            'filename': txt_filename,
            'interesting_findings': interesting_findings.strip(),
            'plugin/theme': current_item,
            'vuln_count': current_vuln_count,
            'details': current_details.strip()
        })

    return results


def saveToExcel(txt_filename, results, workbook, output_file) -> bool:
    """Save processed results to an Excel file."""
    try:
        sheet = workbook.create_sheet(title=os.path.splitext(os.path.basename(txt_filename))[0])
        with open(txt_filename, 'r', encoding='utf-8') as file:
            for i, line in enumerate(file, 1):
                sanitized_line = sanitize_string(line.strip())
                sheet.cell(row=i, column=1, value=sanitized_line)
        if 'Results' not in workbook.sheetnames:
            results_sheet = workbook.create_sheet('Results')
            headers = ["Filename", "Plugin/Theme", "Vuln Count", "Details", "Interesting Findings"]
            results_sheet.append(headers)
        else:
            results_sheet = workbook['Results']
        for result in results:
            row = [
                result.get('filename', ''),
                result.get('plugin/theme', ''),
                result.get('vuln_count', ''),
                result.get('details', ''),
                result.get('interesting_findings', '')
            ]
            results_sheet.append(row)
        apply_formatting(results_sheet)
        workbook.save(output_file)
        return True
    except Exception as e:
        logger.error(f"Error in saveToExcel: {e}")
        return False


def createTwisterResults(input_dir, output_file) -> bool:
    try:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove the default sheet
        all_data = []
        results_sheet = workbook.create_sheet(title='All Results')
        headers = ['Type', 'Domain', 'IP Address', 'NS', 'MX', 'HTTP', 'SMTP', 'ESMTP', 'PHASH']
        results_sheet.append(headers)
        for root, _, files in os.walk(input_dir):
            logger.info(f'Processing files in {root}')
            for file in files:
                if file.endswith('.csv'):
                    with open(os.path.join(root, file), 'r', encoding='utf-8') as f:
                        reader = csv.reader(f)
                        data = list(reader)
                        all_data.extend(data)  # Append data for All Results sheet
                        sheet = workbook.create_sheet(title=os.path.splitext(file)[0])
                        logger.info(f'Data read from {file}')
                        for row in data:
                            sheet.append(row)
        for row in all_data:
            results_sheet.append(row)
        workbook.save(output_file)
        return True
    except Exception as e:
        logger.error(f"Error in createTwisterResults: {e}")
        return False


def main() -> None:
    TODAYIS = datetime.datetime.now().strftime('%d-%m-%y')
    OUTDIR = 'Excel'
    TEMPDIR = 'Excel/temp'
    TWISTERDIR = 'Excel/DNSTwist'
    BACKUPDIR = f'Excel/backup_{TODAYIS}'
    LOGSBACKUPDIR = f'Excel/backup_{TODAYIS}/logs'
    DNSTWISTDIR = 'output/dnstwist'
    WPScan_EXCEL = 'WPScanResults_' + TODAYIS + '.xlsx'
    DNSTwist_EXCEL = 'DNSTwistResults_' + TODAYIS + '.xlsx'
    OUTFILE = os.path.join(OUTDIR, WPScan_EXCEL)
    TWISTEROUTFILE = os.path.join(TWISTERDIR, DNSTwist_EXCEL)
    INDIR = 'logs'
    try:
        setupDirs(OUTDIR, TEMPDIR, BACKUPDIR, TWISTERDIR, LOGSBACKUPDIR, DNSTWISTDIR)
        WPScan_EXCEL = getUniqueFilename(OUTFILE) if os.path.exists(OUTFILE) else OUTFILE
        logger.info(f"Output file is set to {WPScan_EXCEL}")
        if not os.path.exists(INDIR):
            logger.error(f"Input directory {INDIR} does not exist")
            exit()
        copyRecurse(INDIR, TEMPDIR, extension='.log')
        try:
            cleanLogs(TEMPDIR)
            logger.info("Deleted lines before update")
        except Exception as e:
            logger.error(f"Error in delete_lines_before_update: {e}")
        copyRecurse(TEMPDIR, OUTDIR, extension='.txt')
        delAll(TEMPDIR, '.txt')
        delAll(TEMPDIR, '.log')
        if not os.listdir(TEMPDIR):
            logger.info("TEMPDIR is empty")
        workbook = getWorkbook(WPScan_EXCEL)
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']
        if 'Results' not in workbook.sheetnames:
            results_sheet = workbook.create_sheet('Results')
            headers = ["Filename", "Plugin/Theme", "Vuln Count", "Details", "Interesting Findings"]
            results_sheet.append(headers)
        else:
            results_sheet = workbook['Results']
            logger.info("'Results' sheet already exists")
        text_files_processed = False
        try:
            for txt_file in getAllTxtInDir(OUTDIR):
                try:
                    get_results = processFileForResults(txt_file)
                    saveToExcel(txt_filename=txt_file, results=get_results, workbook=workbook, output_file=WPScan_EXCEL)
                    text_files_processed = True
                except Exception as e:
                    logger.error(f"Error processing file {txt_file}: {e}", stack_info=True, exc_info=True, extra={'file': txt_file})
            if not text_files_processed:
                logger.error('No text files found in Excel')
            if text_files_processed:
                logger.info(f"Updated Excel file {WPScan_EXCEL}")
                if not os.listdir(TEMPDIR):
                    pass
                else:
                    pass
                    # logger.error(f"TEMPDIR is not empty: {os.listdir(TEMPDIR)}")
                    # delAll(TEMPDIR, '.txt')
                    # delAll(TEMPDIR, '.log')
        except Exception as e:
            logger.error(f"Error in main function: {e}")
    except Exception as e:
        logger.error(f"Error in main function: {e}")

    # ! Twister section
    try:
        moveLogsToBackup(INDIR, LOGSBACKUPDIR)
        moveLogsToBackup(OUTDIR, BACKUPDIR)
        copyCsvFiles(DNSTWISTDIR, INDIR)
        copyRecurse(DNSTWISTDIR, BACKUPDIR, extension='.png')
        createTwisterResults(input_dir=INDIR, output_file=TWISTEROUTFILE)
        copyRecurse(DNSTWISTDIR, BACKUPDIR, extension='.csv')
        delAll(INDIR, '.csv')
    except Exception as e:
        logger.error(f"Error in main: {e}")


if __name__ == "__main__":
    main()
